from selenium import webdriver
from selenium_stealth import stealth
from seleniumwire import webdriver
from selenium.webdriver.common.by import By
import time
from fake_useragent import UserAgent
import os
from functions import *
import base64
import asyncio
import threading
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import load_workbook
from collections import OrderedDict

cur_dir = os.path.dirname(os.path.abspath(__file__)) 

def fetch(id, auto, loop_string):
    global cur_dir
    wb = load_workbook(cur_dir+'/profiles.xlsx')
    first_sheet = wb.sheetnames[0]
    sheet = wb[first_sheet]  
    
    def pr(text):
        print(f'{id}: {text}')
    

    def replace_vars(sheet, id, str):
        
        for i in range(0, len(sheet[1])):

            f = str.find(f"${sheet[1][i].value}")
            if f != -1:
                id_i = int(id) + 1
                var = sheet[id_i][i].value
                str = str.replace(f"${sheet[1][i].value}", var)

        return(str)
    

    pr(f"запущен профиль")
    
    if auto == True:

        proxy = sheet['B'+str(int(id)+1)].value

        proxy_lp = proxy.split('//')[1].split('@')[0].split(':')
        proxy_ip = proxy.split('//')[1].split('@')[1].split(':')
        proxy_type = proxy.split('//')[0].split(':')[0]

        seleniumwire_options = {
            'proxy': {
                f'{proxy_type}': f'{proxy_type}://{proxy_lp[0]}:{proxy_lp[1]}@{proxy_ip[0]}:{proxy_ip[1]}',
                'verify_ssl': False,
            },
        }

        options = webdriver.ChromeOptions()

        #options.add_argument("start-maximized")
        #options.add_argument("--headless")


        options.add_extension(cur_dir+'/Ex/MetaMask.crx')
        options.add_argument(r"--user-data-dir="+cur_dir+"/Profiles/"+id)
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)

        driver = webdriver.Chrome(options=options ,seleniumwire_options=seleniumwire_options)

        ua = UserAgent()
        fua = ua.chrome

        if not os.path.exists(cur_dir+'/Profiles/'+id):
            os.makedirs(cur_dir+'/Profiles/'+id)
            
        if not os.path.exists(cur_dir+'/Profiles/'+id+'/ua.txt'):
            with open(cur_dir+'/Profiles/'+id+'/ua.txt', 'w') as file:
                file.write(fua)
                p_ua = fua
        else:
            with open(cur_dir+'/Profiles/'+id+'/ua.txt', 'r') as file:
                p_ua = file.read()

        stealth(driver,
                user_agent= p_ua,
                languages=["en-US", "en"],
                vendor="Google Inc.",
                platform="Win32",
                webgl_vendor="Intel Inc.",
                renderer="Intel Iris OpenGL Engine",
                fix_hairline=True,
                )

        
        

        ##########################################################

        loop_string = replace_vars(sheet, id, loop_string)
        
        l_s = loop_string.split(';;')
        
        for s in l_s:
            s = s.strip()
            
            s = s.split('::')
            
            minus = s[0].find('-')
            if minus != -1:
                continue
            
            if s[0] == 'sleep':
                pr(f"Сплю {s[1]} сек.")
                time.sleep(int(s[1]))
            
            if s[0] == 'url':
                try:
                    driver.get(s[1])
                    pr(f"Переход на страницу: {s[1]}")
                except:
                    pr(f'Не удалось перейти на страницу: {s[1]}')
            
            if s[0] == 'click':
                if s[1] == 'xp':
                    try:
                        wait = WebDriverWait(driver, 10)
                        element = wait.until(EC.visibility_of_element_located((By.XPATH, s[2])))
                        element.click()
                        pr(f"Выполнен клик на элемент: {s[2]}")
                    except:
                        pr(f"Не выполнен клик на элемент: {s[2]}")

            if s[0] == 'text':
                if s[1] == 'xp':
                    try:
                        wait = WebDriverWait(driver, 10)
                        element = wait.until(EC.visibility_of_element_located((By.XPATH, s[2])))
                        element.send_keys(s[3])
                        pr(f"Вставлено значение '{s[3]}' в поле: {s[2]}")
                    except:
                        pr(f"Не удалось вставить значение '{s[3]}' в поле: {s[2]}")
        
        
        time.sleep(10)
        pr(f"закрыт профиль")
        driver.quit()
    else:
        print('not_auto')





def worker(semaphore, task_args):
    with semaphore:
        fetch(*task_args)
        
def profiles_range(inp, all):
    try:
        result = []
        ranges = inp.split(',')
        
        for r in ranges:
            if '-' in r:
                start, end = map(int, r.split('-'))
                result.extend(range(start, end + 1))
            else:
                result.append(int(r))
        
        result = list(OrderedDict.fromkeys(result))
        
        for r in result:
            if r > all:
                return False
        return result
        
    except:
        return False
    

def main():
    wb = load_workbook(cur_dir+'/profiles.xlsx')
    first_sheet = wb.sheetnames[0]
    sheet = wb[first_sheet]  
    
    print(f"Найдено профилей: {len(sheet['A'])-1}")
        
    
    mode = int(input('Выберите режим: 1 - отрыть профиля, 2 - выполнить скрипт на профилях:'))
    
    if mode == 1:
        print('ok1')
    elif mode == 2:
        
        pr_r = input('Какие профиля открыть? (1,2,3,5-10...):')
        pr_r = profiles_range(pr_r, len(sheet['A'])-1)
        if pr_r:
            
            pr_scr = input('Введите название скрипта (Пример: test.txt):')
            if os.path.exists(cur_dir+'/Actions/'+pr_scr):
                
                with open(cur_dir+'/Actions/'+pr_scr, 'r') as file:
                    action = file.read()
                
                r_th = int(input('Сколько потоков запустить? (число):'))
                if r_th:
                    
                    num_threads = r_th
                    
                    task_args_list = []
                    for pr in pr_r:
                        task_args_list.append((str(pr), True, action))

                    semaphore = threading.Semaphore(num_threads)

                    threads = []
                    for task_args in task_args_list:
                        thread = threading.Thread(target=worker, args=(semaphore, task_args))
                        thread.start()
                        threads.append(thread)

                    for thread in threads:
                        thread.join()
                
                else:
                    print('Неверное значение')
                
            else:
                print('Файл не найден')
            
        else:
            print('Неверное значение или введены несуществующие профиля')
        
    else:
        print('Неверное значение')
    
    num_threads = 10

    

if __name__ == "__main__":
    main()






# raw_cookies = """
# .twitter.com	TRUE	/	FALSE	1702384887	d_prefs	MToxLGNvbnNlbnRfdmVyc2lvbjoyLHRleHRfdmVyc2lvbjoxMDAw
# .twitter.com	TRUE	/	FALSE	1721047287	guest_id_ads	v1%3A168677764028456149
# .twitter.com	TRUE	/	FALSE	1721047287	guest_id_marketing	v1%3A168677764028456149
# .twitter.com	TRUE	/	FALSE	1721047287	personalization_id	"v1_XcL0WIIELaF9ZwYwRalRdw=="
# .twitter.com	TRUE	/	FALSE	1722062155	guest_id	v1%3A168784775518531926
# twitter.com	FALSE	/	FALSE	1703399795	g_state	{"i_l":0}
# .twitter.com	TRUE	/	FALSE	1722062195	kdt	L4z05oovtqJF2nO2Ig21kr34vQbYQJB3P0G4zqqa
# .twitter.com	TRUE	/	FALSE	1722062195	auth_token	c8e7c538bc8c4e561d1feea1ad697ac09bcc620a
# .twitter.com	TRUE	/	FALSE	1722375467	dnt	1
# .twitter.com	TRUE	/	FALSE	1726299110	ct0	6e590ae3ac52893c402ece7693868ae0a8aeec75415f20903d553bbc2127aeeffe88f19ff98939c740b7aa9ae5529fe300cd5fd4ae571c27ccc02972d64a9b9b249cb1573e339aff2d5d0d751b8c8b59
# .twitter.com	TRUE	/	FALSE	1723931783	twid	u%3D1492814130281074688
# .twitter.com	TRUE	/	FALSE	1693000583	external_referer	padhuUp37zjgzgv1mFWxJ12Ozwit7owX|0|8e8t2xd8A2w%3D
# .twitter.com	TRUE	/	FALSE	1755467783	_ga	GA1.2.278601459.1686832891
# .twitter.com	TRUE	/	FALSE	1692482183	_gid	GA1.2.1375315745.1692395784
# """

# cookie_lines = raw_cookies.strip().split('\n')

# for line in cookie_lines:
#     parts = line.split('\t')
#     cookie = {
#         'name': parts[5],
#         'value': parts[6],
#         'domain': parts[0],
#         'path': parts[2],
#         'secure': parts[3] == 'TRUE',
#         'expiry': int(parts[4])
#     }
#     driver.add_cookie(cookie)
 



# base64_cookies = [
#     "LS50LmNvbS9UV0FMUkRPUy9GVUxBU0UgMTcwMjM4NDg4NyBkX3ByZWZzIE1Ub3hMTkdNd3k6MTYzNzg4MjA=",
#     "LS50LmNvbS9UV0FMUkRPUy9GVUxBU0UgMTcyMTA0Nzg3IGd1ZXN0X2lkX2FkcyB2MSUzQTE2ODY3Nzc2NDAyODQ1NjE0OQ==",
#     # Добавьте другие куки в формате base64 по аналогии
# ]

# # Перебираем куки и добавляем их к драйверу
# for base64_cookie in base64_cookies:
#     cookie_data = base64.b64decode(base64_cookie).decode("utf-8")
#     parts = cookie_data.split("\t")
#     cookie = {
#         'name': parts[5],
#         'value': parts[6],
#         'domain': parts[0],
#         'path': parts[2],
#         'secure': parts[3] == 'TRUE',
#         'expiry': int(parts[4])
#     }
#     driver.add_cookie(cookie)




# cookie = {
#     'name': 'auth_token',
#     'value': '94efe8def2ad9fc038b62ec808d9e209b02247a9',
#     'domain': '.twitter.com',  # Важно: укажите поддомен, начинающийся с точки
#     'path': '/',
#     'secure': False,  # True, если используется HTTPS
#     'expiry': 9999999999  # Время истечения куки (в секундах с начала эпохи)
# }

# a = """
# b5d4f5bb9e1960c1a6df539366b9c3eaa5fc278f
# b8f073a9874cf3d9ff1544f72cd576dd73c1e53a
# eb7dc55aaf30df3194c5a9c9e49f002876523353
# 7f1535cd0afcbed670a4ecf06ca9b6ac23c2123b
# a0fa2212f3212c1a488de0c3e5c8739e9be05aa7
# cbcb9617aa6c50d748e9679ad149fb4bd0cd600e
# 93e1bcc36334457418cd441d3515e9f4234d506d
# 0e6d966bb3ee0017016722ad1513860ff70ee726
# 0a924006f4f40df6392fc617331d0a92901f4ba4
# a4c2b88bc4d3b13457f2b72723e8ea7de7dd926b
# 998830c50f6629666b7877630e99eccc64e05b19
# 92fec745e37275f3629bd792a0536c709783426b
# 33cc2fa2fd0fc5b534b6954a76f961de685d96ac
# 8425e11bf968bb4c952a6ae53f33d427c5dffc96
# 2c47e548d9e4668fce931a6a5e35832ba7cb16cf
# 064b7ba2520c84f9cb16d8e767160fdccc7e498c
# 33f52aea2dedc9b2eb32fad9621b8b9af2b30062
# f77df997e9d5f5f436c9a8da85fadfe4d0d8fa53
# a4bd6c53ddc8040648dc03de45081e4f3c21353d
# 94efe8def2ad9fc038b62ec808d9e209b02247a9
# d2b7a0f289a023f9803f023ab102f22e15ab5048
# 4854c26c6d73b0406211b2e492f7d002c21e10f0
# 0380010a6ef6010e7937afdc526e33ea230db1bb
# 2eded78c5e264a336e17e10294dcee35b44e65b2
# """


# # Добавляем куки к драйверу
# driver.add_cookie(cookie)

# Обновляем страницу после добавления куки
# driver.refresh()




